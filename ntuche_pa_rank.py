import numpy as np
import pandas as pd
import os
import openpyxl
import decimal
from openpyxl.utils.dataframe import dataframe_to_rows

class arrangement:
    
    def __init__(self, studentgrade, path_set, path, core_course2, che_course_number):
        """
        
        初始化
        
        ----------
        Parameters
        ----------
        studentgrade: int
            學生年級
        path : str
            學生成績的檔案路徑
        __df_alldata : pd.DataFrame
            所有學生的所有平均分數資料
        """
        self.studentgrade = studentgrade
        self.path_set = path_set
        self.path = path
        self.core_course2 = core_course2
        self.che_course_number = che_course_number
        self.__df_alldata = None
    
    @staticmethod
    def modify_round(x, dec=2):
        """
        
        四捨五入的函數
        (python的round、numpy的round和around都常常會回傳錯誤的結果, eg. 4.165回傳4.16給我)
        
        ----------
        Parameters
        ----------
        x: float
            想要取四捨五入的數值
        dec: float
            四捨五入的精確度(eg.想要取至小數點下第2位就設為2)
        rounded_x: str/float
            取完四捨五入以後的數值
        """
        x_str = str(x)
        x = decimal.Decimal(x_str)
        rounded_x = x.quantize(decimal.Decimal(str(10**(-dec))), rounding=decimal.ROUND_HALF_UP)
        return float(rounded_x)
    
    @staticmethod
    def dedupe(items):
        """
        
        在不影響順序的情況下刪除重複元素的函數
        
        ----------
        Parameters
        ----------
        items: iterable list, int
            有許多重複元素的list或array
        """
        seen = set()
        for item in items:
            if item not in seen:
                yield item
                seen.add(item)
    
    @property
    def df_setdata(self):
        """
        Parameters
        ----------
        df_setdata: pd.DataFrame
            化工系所有必修課目的列表
        """
        return pd.read_excel(self.path_set,'112_major_subject').iloc[:-1].replace('\xa0\xa0', np.nan) #將空格取代成nan
    
    @property
    def core_course1(self):
        """
        Parameters
        ----------
        core_course1: list, int
            化工系所有必修課目的list
        """
        df_setdata = self.df_setdata
        core_course1 = list(df_setdata['Course Name'])
        return core_course1
    
    @property
    def df_gradedata(self): #學生成績
        """
        Parameters
        ----------
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        col: list, str
            成績總表的欄位名稱
        """
        df_gradedata = pd.read_excel(self.path).replace('\xa0\xa0', np.nan)
        col = [str(i1)+str(i2) if type(i2) is str else i1 for i1, i2 in zip(df_gradedata.iloc[1], df_gradedata.iloc[2])] #整理獲得column名稱
        df_gradedata.columns = col
        df_gradedata = df_gradedata.iloc[3:].reset_index(drop=True) #刪除dataframe內的column名稱
        return df_gradedata
    
    @property
    def all_students_id(self): 
        """
        Parameters
        ----------
        all_students_id: list, str
            所有學生的學號
        """
        return [i.strip() for i in self.dedupe((self.df_gradedata['學號']))]
    
    @property
    def all_students_name(self):
        """
        Parameters
        ----------
        all_students_name: list, str
            所有學生的姓名
        """
        return [i.strip() for i in self.dedupe((self.df_gradedata['姓名']))]
    
    def calc_allavg(self, student_id, full_output=False):
        """
        
        計算一個學生所有科目的總平均(大二與大三的同分參酌一、大四的排名標準)
        
        ----------
        Parameters
        ----------
        student_id: str
            學生的學號
        full_output: boolean
            是否需要輸出學生的總學分數
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        grade: np.array, float
            學生各科的成績(等第積分)
        credit: np.array, int
            學生各科的學分數
        flag: boolean
            是否已搜尋到此學生的資料
        allcredit: int
            學生的總學分數
        allavg: float
            學生所有科目的平均分數
        """
        df_gradedata = self.df_gradedata
        grade = np.array([])
        credit = np.array([])
        flag = False
        for (ide, cde, gde) in zip(df_gradedata['學號'], df_gradedata['學分'], df_gradedata['等第績分']):
            if ide.strip() == student_id and not np.isnan(float(gde)):
                grade = np.append(grade, float(gde.strip()))
                credit = np.append(credit, int(cde.strip()))
                flag = True
            if ide.strip() != student_id and flag:
                break
        allcredit = sum(credit)
        allavg = np.sum( grade * credit ) / allcredit if ( grade.size != 0 ) and ( allcredit != 0 ) else 0
        allavg = self.modify_round(allavg)
        if full_output:
            return allavg, allcredit
        else:
            return allavg
    
    def calc_core1avg(self, student_id, full_output=False):
        """
        
        計算一個學生必修科目總平均(大二與大三的排名標準)
        
        ----------
        Parameters
        ----------
        student_id: str
            學生的學號
        full_output: boolean
            是否需要輸出學生的各個必修課目的課程名稱、等第成績、等第積分與學分數資料
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        core_course1: list, int
            化工系所有必修課目名稱
        core_course1_name: list, str
            學生修習的各個必修課目名稱
        grade: np.array, float
            學生修習的各個必修課目成績(等第積分)
        credit: np.array, int
            學生修習的各個必修課目的學分數
        gdcddata: list, int
            學生修習的各個必修課目的"等第成績 等第積分 學分數"
        flag: boolean
            是否已搜尋到此學生的資料
        core1credit: int
            學生必修課的總學分數
        core1avg: float
            學生必修課的平均分數
        fulldata: dict, tuple, str
            包含學生修習的各個必修課目的課程名稱、等第成績、等第積分與學分數, 等第成績代表A+, A, A-, ...等
        """
        df_gradedata = self.df_gradedata
        core_course1 = self.core_course1
        core_course1_name = []
        grade = np.array([])
        credit = np.array([])
        gdcddata = []
        flag = False
        for (ide, cne, cde, gne, gde) in zip(df_gradedata['學號'], df_gradedata['課程名稱'], df_gradedata['學分'], df_gradedata['等第成績'], df_gradedata['等第績分']):
            if ide.strip() == student_id and not np.isnan(float(gde)) and cne.strip() in core_course1 :
                core_course1_name.append(cne.strip())
                grade = np.append(grade, float(str(gde).strip()))
                credit = np.append(credit, int(cde.strip()))
                gdcddata.append(gne.strip() + ' ' + str(gde).strip() + ' ' + cde.strip())
                flag = True
            if ide.strip() != student_id and flag:
                break
        core1credit = sum(credit)
        core1avg = np.sum( grade * credit ) / core1credit if ( grade.size != 0 ) and ( core1credit != 0 ) else 0
        core1avg = self.modify_round(core1avg)
        if full_output:
            fulldata = dict(zip(core_course1_name, gdcddata))
            return core1avg, fulldata
        else:
            return core1avg
    
    def calc_core2avg(self, student_id):
        """
        
        計算一個學生必修科目核心課程的平均(大二與大三的同分斟酌二)
        
        ----------
        Parameters
        ----------
        student_id: str
            學生的學號
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        core_course2 : list, str
            化工系所有必修課的核心課程科目
        grade: np.array, float
            學生修習的各個必修課的核心課程成績(等第積分)
        credit: np.array, int
            學生修習的各個必修課的核心課程學分數
        flag: boolean
            是否已搜尋到此學生的資料
        core2credit: int
            學生必修課核心課程的總學分數
        core2avg: float
            學生必修課核心課程的平均分數
        """
        df_gradedata = self.df_gradedata
        core_course2 = self.core_course2
        grade = np.array([])
        credit = np.array([])
        flag = False
        for (ide, cne, cde, gde) in zip(df_gradedata['學號'], df_gradedata['課程名稱'], df_gradedata['學分'], df_gradedata['等第績分']):
            if ide.strip() == student_id and not np.isnan(float(gde)) and cne.strip() in core_course2:
                grade = np.append(grade, float(str(gde).strip()))
                credit = np.append(credit, int(cde.strip()))
                flag = True
            if ide.strip() != student_id and flag:
                break
        core2credit = sum(credit)
        core2avg = np.sum( grade * credit ) / core2credit if ( grade.size != 0 ) and ( core2credit != 0 ) else 0
        core2avg = self.modify_round(core2avg)
        return core2avg
    
    def calc_core3avg(self, student_id, full_output=False):
        """
        
        計算一個學生化工系課程的平均分數(大四的同分參酌一)
        
        ----------
        Parameters
        ----------
        student_id: str
            學生的學號
        full_output: boolean
            是否需要輸出學生的各個化工系課程的課程名稱、等第成績、等第積分與學分數資料
        df_gradedata: pd.DataFrame
            所有學生的成績總表
        che_course_number: list, str
            化工系開設課程之課號前三碼
        che_course_name: list, str
            學生修習的各個化工系課程名稱
        grade: np.array, float
            學生修習的各個化工系課程成績(等第積分)
        credit: np.array, int
            學生修習的各個化工系課程的學分數
        gdcddata: list, int
            學生修習的各個化工系課程的"等第成績 等第積分 學分數", 等第成績代表A+, A, A-, ...等
        flag: boolean
            是否已搜尋到此學生的資料
        core3_number: int
            學生修習的化工系課程數目
        core3credit: int
            學生修習化工系課程的總學分數
        core3avg: float
            學生修習化工系課程的平均分數
        fulldata: dict, tuple, str
            包含學生修習的各個化工系課程的課程名稱、等第成績、等第積分與學分數
        """
        df_gradedata = self.df_gradedata
        che_course_number = self.che_course_number
        che_course_name = []
        grade = np.array([])
        credit = np.array([])
        gdcddata = []
        flag = False
        core3_number = 0
        for (ide, cne, cie, cde, gne, gde) in zip(df_gradedata['學號'], df_gradedata['課程名稱'], df_gradedata['課程識別碼'], df_gradedata['學分'], df_gradedata['等第成績'], df_gradedata['等第績分']):
            if ide.strip() == student_id and not np.isnan(float(gde)) and cie.strip()[:3] in che_course_number:
                che_course_name.append(cne.strip())
                grade = np.append(grade, float(str(gde).strip()))
                credit = np.append(credit, int(cde.strip()))
                gdcddata.append(gne.strip() + ' ' + str(gde).strip() + ' ' + cde.strip())
                core3_number += 1
                flag = True
            if ide.strip() != student_id and flag:
                break
        core3credit = np.sum(credit)
        core3avg = np.sum( grade * credit ) / core3credit if ( grade.size != 0 ) and ( core3credit != 0 ) else 0
        core3avg = self.modify_round(core3avg)
        if full_output:
            fulldata = dict(zip(che_course_name, gdcddata))
            return core3avg, core3_number, fulldata
        else:
            return core3avg
    
    def get_df_alldata(self):
        """
        
        計算所有學生的所有平均分數資料
        
        ----------
        Parameters
        ----------
        studentgrade: int
            學生年級
        all_students_id: list, str
            所有學生的學號
        all_students_name: list, str
            所有學生的姓名
        all_allavg: list, float
            所有學生的全科目平均分數
        all_allcredit: list, int
            所有學生的總學分數
        all_core1avg: list, float
            所有學生的必修課平均分數
        all_core2avg: list, float
            所有學生的必修課平均分數
        all_core3avg: list, float
            所有學生的化工系課程平均分數
        all_core3_number: list, float
            所有學生修習的化工系課程數目
        df_corse1data: pd.DataFrame
            大二與大三：所有學生修習各個必修課的等第成績、等第積分與學分數總表
            大四：所有學生修習各個化工系課程的等第成績、等第積分與學分數總表
        df_avgdata: pd.DataFrame
            大二與大三：所有學生的必修課平均分數、全科目平均分數(同分參酌一)、必修課的核心課程平均分數(同分參酌二)與總學分數資料表
            大四：所有學生的全科目平均分數、化工系課程平均分數(同分參酌一)、總學分數(同分參酌二)與修習的化工系課程數目資料表
        column: list, str
            df_avgdata的欄位名稱
        df_alldata: pd.DataFrame
            df_avgdata和df_corse1data合併後的總表
        """
        studentgrade = self.studentgrade
        all_students_id = self.all_students_id
        all_students_name = self.all_students_name
        all_allavg = []
        all_allcredit = []
        df_corse1data = pd.DataFrame()
        if studentgrade in [2, 3]:
            all_core1avg = []
            all_core2avg = []
            column = ['學號','姓名','必修平均','同分參酌一','同分參酌二', '總學分數']
            for student_id in all_students_id:
                allavg, allcredit = self.calc_allavg(student_id, True)
                all_allavg.append(allavg)
                all_allcredit.append(allcredit)
                core1avg, fulldata = self.calc_core1avg(student_id, True)
                df_corse1data = pd.concat([df_corse1data, pd.DataFrame([fulldata])], ignore_index=True)
                all_core1avg.append(core1avg)
                all_core2avg.append(self.calc_core2avg(student_id))
            df_avgdata = pd.DataFrame(zip(all_students_id, all_students_name, all_core1avg, all_allavg, all_core2avg, all_allcredit),columns=column)
        elif studentgrade == 4:
            all_core3avg = []
            all_core3_number =[]
            column = ['學號','姓名','所有科目平均','同分參酌一','同分參酌二', '修習化工系課程數目']
            for student_id in all_students_id:
                allavg, allcredit = self.calc_allavg(student_id, True)
                all_allavg.append(allavg)
                all_allcredit.append(allcredit)
                core3avg, core3_number, fulldata = self.calc_core3avg(student_id, True)
                df_corse1data = pd.concat([df_corse1data, pd.DataFrame([fulldata])], ignore_index=True)
                all_core3avg.append(core3avg)
                all_core3_number.append(core3_number)
            df_avgdata = pd.DataFrame(zip(all_students_id, all_students_name, all_allavg, all_core3avg, all_allcredit, all_core3_number),columns=column)
        df_alldata = pd.concat([df_avgdata, df_corse1data], axis=1)
        self.__df_alldata = df_alldata
        return df_alldata
    
    @property
    def df_alldata(self):
        """
        Parameters
        ----------
        df_alldata: pd.DataFrame
            所有學生的所有平均分數資料總表
        """
        if self.__df_alldata is None:
            return self.get_df_alldata()
        else:
            return self.__df_alldata
    
    @property
    def df_rankdata(self):
        """
        
        進行排名的計算
        
        ----------
        Parameters
        ----------
        studentgrade: int
            學生年級
        col_all: list, str
            排名所依照的各種先後順序，順序由左到右
        df_alldata: pd.DataFrame
            所有學生的所有平均分數資料總表
        df_rankdata: pd.DataFrame
            包含所有學生所有平均分數資料的排名總表
        """
        studentgrade = self.studentgrade
        if studentgrade in [2, 3]:
            col_all = ['必修平均','同分參酌一','同分參酌二']
        elif studentgrade == 4:
            col_all = ['所有科目平均','同分參酌一','同分參酌二']
        df_alldata = self.df_alldata
        df_rankdata = df_alldata.copy()
        ranklist = df_rankdata[col_all].apply(tuple, axis=1).rank(method='min', ascending=0)
        df_rankdata.insert(0, '排名', ranklist) #插入一欄紀錄每位學生的排名
        df_rankdata.sort_values('排名', inplace=True) #將此表格以排名來排序
        df_rankdata.index = df_alldata.index
        return df_rankdata
    
    def save_rankdata(self, savepath, sheet_name, method='dataframe_to_rows'):
        """
        
        將排名後的資料儲存至指定路徑
        
        ----------
        Parameters
        ----------
        savepath: str
            排名結果的檔案儲存路徑
        sheet_name: str
            設定結果的excel檔中的工作表名稱
        method: str
            存檔的方法有兩種:
            1. ExcelWriter: 程式碼比較簡潔，但我當初在編寫時有時候會產生出損毀過的excel檔
            2. dataframe_to_rows: 程式碼看起來比較繁雜，但是可以產生出正常的excel檔
        df_rankdata: pd.DataFrame
            包含所有學生所有平均分數資料的排名總表
        """
        df_rankdata = self.df_rankdata
        if method == 'ExcelWriter':
            if os.path.exists(savepath):
                writer = pd.ExcelWriter(savepath, engine='openpyxl', mode='a')
                book = openpyxl.load_workbook(savepath)
                writer.book = book
            else:
                writer = pd.ExcelWriter(savepath, engine='openpyxl')
                book = openpyxl.Workbook()
            df_rankdata.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
            writer.close()
        elif method == 'dataframe_to_rows':
            rows = dataframe_to_rows(df_rankdata, index=False)
            if os.path.exists(savepath):
                book = openpyxl.load_workbook(savepath)
                sheet = book.create_sheet(title=sheet_name)
            else:
                book = openpyxl.Workbook()
                sheet = book.active
                sheet.title = sheet_name
            for r_idx, row in enumerate(rows, 1):
                for c_idx, value in enumerate(row, 1):
                     sheet.cell(row=r_idx, column=c_idx, value=value)
            book.save(filename=savepath)
        else:
            print('Please input "dataframe_to_rows" or "ExcelWriter" to method variable.')